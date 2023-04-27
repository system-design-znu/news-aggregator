import asyncio
import requests
import aiohttp
import logging
import re
import asyncpg
import openpyxl
import pathlib
import hashlib
import time
import configparser
from aio_pika import (
    connect as rabbit_mq_connector,
    Message as rabbit_mq_message,
    DeliveryMode as rabbit_mq_delivery_mode,
    ExchangeType as rabbit_mq_exchange_type,
)
from aio_pika.exceptions import ChannelClosed as rabbit_mq_does_not_exists_exception
import time
from signal import signal, SIGINT, SIG_DFL

config = configparser.ConfigParser()
config.read(str(pathlib.Path(__file__).parent.resolve()) + "/config.ini")
db_user = config["postgres"]["user"]
db_pass = config["postgres"]["pass"]
db_port = config["postgres"]["port"]
db_table_name = config["postgres"]["table_name"]
rabbit_mq_host = 'rabbit_mq'
# rabbit_mq_host, postgres_host = ["localhost" for x in range(2)]
postgres_host = 'postgres_db'
database_conn = None
old_excel_file_hash = None
excel_file_path = str(pathlib.Path(__file__).parent.resolve()) + "/SQL/RSS.xlsx"
(
    initializer_finish_event,
    change_excel_event,
    shutdown_event,
    update_rss_container_activation_event,
    generator_finish_work_event,
    populate_domain_queue_event,
    domains_queue_container_update_event,
) = [asyncio.Event() for i in range(7)]
generator_wants_to_run = 0
rabbit_mq_conn = None
shutdown_command_issued = False
# is_running_in_docker = True if config['docker']['is_running_in_docker'] == 'True' else False
# localhost_addr = 'host.docker.internal' if is_running_in_docker else 'localhost'
available_workers, rss_container = list(), list()
logging.basicConfig(
    format="%(asctime)s - %(levelname)s:  %(message)s",
    datefmt="%d-%b-%y %H:%M:%S",
    level=logging.INFO,
)


async def initializer():
    global rabbit_mq_conn
    global database_conn
    while True:
        try:
            rabbit_mq_conn = await rabbit_mq_connector(
                f"amqp://guest:guest@{rabbit_mq_host}:5672/"
            )
        except:
            continue
        else:
            break
    try:
        database_conn = await asyncpg.connect(
            user=db_user,
            password=db_pass,
            host=postgres_host,
            port=db_port,
            database=db_table_name,
        )
    except asyncpg.exceptions.InvalidCatalogNameError:
        conn = await asyncpg.connect(user=db_user, password=db_pass, host=postgres_host)
        values = await conn.execute(
            f"""
                CREATE DATABASE {db_table_name};
            """
        )
        await conn.close()
        database_conn = await asyncpg.connect(
            user=db_user,
            password=db_pass,
            host=postgres_host,
            port=db_port,
            database=db_table_name,
        )
    logging.info("Initialization finished.")
    initializer_finish_event.set()


async def excel_file_hash_checker():
    await initializer_finish_event.wait()
    global old_excel_file_hash
    while True:
        md5 = hashlib.md5()
        BUF_SIZE = 65536
        with open(excel_file_path, "rb") as f:
            start_time = time.time()
            while True:
                data = f.read(BUF_SIZE)
                if not data:
                    break
                md5.update(data)
                if time.time() - start_time > 0.1:
                    await asyncio.sleep(0)

        if (
            old_excel_file_hash == None
            or old_excel_file_hash.hexdigest() != md5.hexdigest()
        ):
            logging.info(
                "An update for the RSS urls is required because the hash of excel file containing the urls changed recently."
            )
            old_excel_file_hash = md5
            change_excel_event.set()
        await asyncio.sleep(1)


def get_data_from_excel():
    dataframe = openpyxl.load_workbook(excel_file_path)
    dataframe1 = dataframe.active
    rows_to_write = list()
    for index, row in enumerate(range(0, dataframe1.max_row)):
        if index == 0:
            continue
        temp_list = list()
        for index, col in enumerate(dataframe1.iter_cols(1, dataframe1.max_column)):
            if index in [1, 5]:
                if index == 1:
                    temp_list.append(col[row].value)
                elif index == 5:
                    temp_list.append(float(col[row].value))
        rows_to_write.append(temp_list)
    return rows_to_write


async def update_urls_table():
    await initializer_finish_event.wait()
    while True:
        await change_excel_event.wait()
        if not update_rss_container_activation_event.is_set():
            logging.info("RSS table update is in progress...")
            await database_conn.execute(
                """
                    DROP TABLE IF EXISTS public.rss_fetch;
                    CREATE TABLE public.rss_fetch (
                    url varchar NOT NULL,
                    priority varchar NOT NULL);
                """
            )
            rows_to_write = await asyncio.get_running_loop().run_in_executor(
                None, get_data_from_excel
            )
            await database_conn.execute(
                """
                                            INSERT INTO public.rss_fetch (url, priority)
                                            VALUES {}
                                        """.format(
                    "".join(["('{}', '{}'),".format(i, j) for i, j in rows_to_write])[
                        :-1
                    ]
                    + ";"
                )
            )
            change_excel_event.clear()
            update_rss_container_activation_event.set()
            logging.info("RSS table update finished.")
        await asyncio.sleep(0.1)


def get_workers_list(
    user="guest", password="guest", host=rabbit_mq_host, port=15672, virtual_host=None
):
    url = "http://%s:%s/api/queues/%s" % (host, port, virtual_host or "")
    response = requests.get(url, auth=(user, password))
    queues = [q["name"] for q in response.json()]
    return queues


async def get_workers_list_async(
    user="guest", password="guest", host=rabbit_mq_host, port=15672, virtual_host=None
):
    async with aiohttp.ClientSession() as session:
        url = "http://%s:%s/api/queues/%s" % (host, port, virtual_host or "")
        async with session.get(url, auth=aiohttp.BasicAuth(user, password)) as response:
            workers = [
                worker["name"]
                for worker in await response.json()
                if "fetch_worker" in worker["name"]
            ]
            return workers


async def add_fetch_worker(
    user="guest", password="guest", host=rabbit_mq_host, port=15672, virtual_host=None
):
    await initializer_finish_event.wait()
    global available_workers
    url = "http://%s:%s/api/queues/%s" % (host, port, virtual_host or "")
    while True:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url, auth=aiohttp.BasicAuth(user, password)
            ) as response:
                result = await response.json()
                workers = [worker["name"] for worker in result]
                for worker in workers:
                    if "fetch_worker" in worker:
                        if worker not in available_workers:
                            logging.info(
                                f'A new fetch worker joined to the system: "{worker}"'
                            )
                            available_workers.append(worker)
        await asyncio.sleep(3)


async def rss_container_updater():
    await initializer_finish_event.wait()
    global rss_container
    while True:
        await update_rss_container_activation_event.wait()
        rss_container = list()
        received_rss_data_from_database = await database_conn.fetch(
            """
                SELECT url, priority
                FROM public.rss_fetch;
            """
        )
        # print(received_rss_data_from_database)
        for record in received_rss_data_from_database:
            temp_list = []
            for value in record.values():
                temp_list.append(value)
            rss_container.append(
                {
                    "url": temp_list[0],
                    "priority": float(temp_list[1]),
                    "last_process_time": 0,
                }
            )
        rss_container.sort(key=lambda x: x["priority"])
        update_rss_container_activation_event.clear()
        await asyncio.sleep(0.1)


async def gracefully_async_exit_handler():
    await shutdown_event.wait()


def gracefully_sync_exit_handler(sig, frame):
    """Closes the database and stops the generator gracefully.

    Args:
        sig (int): The signal id this handler is supposed to handle. In our
        case, it is the interrupt signal (2).

        frame (frame): The stack frame is stored in this variable. We don't need
        to use it.
    """
    # TODO: Close database connection and stop generator
    global shutdown_command_issued, shutdown_event
    shutdown_command_issued = True
    asyncio.run(gracefully_async_exit_handler())
    quit()


pattern = r"^(?:https?:\/\/)?(?:[^@\/\n]+@)?(?:www\.)?([^:\/\n]+)"
# # Handling the interrupt signal.
# signal(SIGINT, gracefully_sync_exit_handler)
domains_queue_container = dict()


async def add_new_urls_to_domain_queue():
    await initializer_finish_event.wait()
    global domains_queue_container
    # domains_queue_container = {
    #     domain_name: {
    #         priority1: ['ready to fetch url 1', 'ready to fetch url 2'], # keys are priorities
    #         priority2: ['ready to fetch url 1', 'ready to fetch url 2']
    #     }
    # }
    while True:
        if (
            not domains_queue_container_update_event.is_set()
            or not populate_domain_queue_event.is_set()
        ):
            for item in rss_container:
                if (
                    time.time() - item["last_process_time"] > 900
                    if item["priority"] == 1
                    else 3600
                    if item["priority"] == 2
                    else 7200
                ):
                    extracted_domain = re.search(pattern, item["url"]).group(1)
                    if extracted_domain not in domains_queue_container.keys():
                        domains_queue_container.update({extracted_domain: dict()})
                    if (
                        item["priority"]
                        not in domains_queue_container[extracted_domain].keys()
                    ):
                        domains_queue_container[extracted_domain].update(
                            {item["priority"]: list()}
                        )
                    if "fetch_counts" not in domains_queue_container[extracted_domain]:
                        domains_queue_container[extracted_domain][
                            "fetch_counts"
                        ] = dict()
                    if (
                        item["url"]
                        not in domains_queue_container[extracted_domain][
                            item["priority"]
                        ]
                    ):
                        domains_queue_container[extracted_domain][
                            item["priority"]
                        ].append(item["url"])
                    if (
                        item["priority"]
                        not in domains_queue_container[extracted_domain]["fetch_counts"]
                    ):
                        domains_queue_container[extracted_domain]["fetch_counts"][
                            item["priority"]
                        ] = 1
                    await asyncio.sleep(0)
                    populate_domain_queue_event.set()
            if generator_wants_to_run:
                domains_queue_container_update_event.set()
        await asyncio.sleep(5)


async def generator():
    await initializer_finish_event.wait()
    await populate_domain_queue_event.wait()

    """The while loop in this function will be executed in certain times
    and sends the processable URLs to a RabbitMQ exchange to be received by
    "URL Fetching Unit".
    """
    sleep_time = 0.01
    global shutdown_command_issued, shutdown_event, available_workers, rss_container, domains_queue_container, generator_wants_to_run
    async with rabbit_mq_conn:
        rabbit_mq_channel = await rabbit_mq_conn.channel()
        try:
            # Checking whether the exchange we want to use exists. If not, an
            # error will be thrown by RabbitMQ. We catch that, refresh the
            # channel (because it will be closed when an exception occurs) and
            # create a new exchange.
            await rabbit_mq_channel.declare_exchange(
                "news_aggregator_direct",
                rabbit_mq_exchange_type.DIRECT,
                durable=True,
                passive=True,
            )
        except rabbit_mq_does_not_exists_exception:
            rabbit_mq_channel = await rabbit_mq_conn.channel()
            rabbit_mq_direct_exchange = await rabbit_mq_channel.declare_exchange(
                "news_aggregator_direct", rabbit_mq_exchange_type.DIRECT, durable=True
            )
            logging.info("A new rabbit mq exchange created: news_aggregator_direct")
        else:
            rabbit_mq_direct_exchange = await rabbit_mq_channel.declare_exchange(
                "news_aggregator_direct", rabbit_mq_exchange_type.DIRECT, durable=True
            )
        # available_workers = { worker: { 'domain': 'left_fetch' } }
        while True:
            worker_disconnect = False
            if not shutdown_command_issued:
                shutdown_command_issued = False
                # If the rss container is not under an update
                if not update_rss_container_activation_event.is_set():
                    generator_wants_to_run = True
                    await domains_queue_container_update_event.wait()
                    if len(rss_container) != 0 and len(available_workers) != 0:
                        worker_item_index = 0
                        temp_available_workers = available_workers.copy()
                        print('this is the temp available workers', temp_available_workers)
                        while worker_item_index < len(available_workers):
                            current_worker = available_workers[worker_item_index]
                            if not available_workers[worker_item_index] in await get_workers_list_async():
                                del temp_available_workers[worker_item_index]
                                logging.error(f'The worker "{current_worker}" has been disconnected.')
                                worker_disconnect = True
                                worker_item_index += 1
                                continue
                            for domain, domain_info in domains_queue_container.items():
                                left_fetches = 5
                                for (
                                    priority,
                                    ready_to_fetch_urls,
                                ) in domain_info.items():
                                    if priority == "fetch_counts":
                                        continue
                                    jump_to_next_priority = None
                                    while (
                                        len(domains_queue_container[domain][priority])
                                        > 0
                                    ):
                                        if (
                                            priority + 1
                                            in domains_queue_container[domain][
                                                "fetch_counts"
                                            ]
                                            and (
                                                domains_queue_container[domain][
                                                    "fetch_counts"
                                                ][priority]
                                                / domains_queue_container[domain][
                                                    "fetch_counts"
                                                ][priority + 1]
                                            )
                                            > 2
                                            and len(
                                                domains_queue_container[domain][
                                                    priority + 1
                                                ]
                                            )
                                            > 0
                                        ):
                                            jump_to_next_priority = True
                                            break
                                        elif left_fetches > 0:
                                            url = domains_queue_container[domain][
                                                priority
                                            ].pop(0)
                                            try:
                                                await rabbit_mq_direct_exchange.publish(
                                                    rabbit_mq_message(
                                                        bytes(url, encoding="utf-8"),
                                                        delivery_mode=rabbit_mq_delivery_mode.PERSISTENT,
                                                    ),
                                                    routing_key=current_worker,
                                                )
                                                logging.info(
                                                f"The URL \"{url}\" that has a priority of {priority}, will be fetched by the worker {current_worker} soon."
                                                )
                                            except Exception as e:
                                                logging.critical('Something went wrong during sending data to RabbitMQ. The generator unit will be terminated in order to be fixed later.')
                                                print(e)
                                            domains_queue_container[domain][
                                                "fetch_counts"
                                            ][priority] += 1
                                            left_fetches -= 1
                                        else:
                                            break
                                        if jump_to_next_priority:
                                            break
                            worker_item_index += 1
                        worker_item_index = 0
                        # We will visit each domain every 15 mins (900 seconds) and
                        # will fetch 5 times from that domain on every visit (which occurs every 900 seconds).
                        domains_queue_container_update_event.clear()
                        generator_wants_to_run = False
                        if worker_disconnect:
                            available_workers = temp_available_workers
                        await asyncio.sleep(900)
                    else:
                        await asyncio.sleep(sleep_time)
                else:
                    await asyncio.sleep(sleep_time)
            else:
                break
    logging.info("The connection with the RabbitMQ node closed.")
    shutdown_event.set()


loop = asyncio.new_event_loop()
loop.create_task(initializer())
loop.create_task(excel_file_hash_checker())
loop.create_task(update_urls_table())
loop.create_task(rss_container_updater())
loop.create_task(generator())
loop.create_task(add_fetch_worker())
loop.create_task(add_new_urls_to_domain_queue())
loop.run_forever()
